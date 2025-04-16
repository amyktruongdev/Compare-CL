import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.cell.cell import Cell
import matplotlib.pyplot as plt
import seaborn as sns

st.set_page_config(page_title="CL Comparison Tool", layout="centered")
st.title("📊 CL Comparison Tool")

num_files = st.selectbox("Select number of files to compare", [2, 3, 4], index=0)

uploaded_files = []
for i in range(num_files):
    uploaded_file = st.file_uploader(f"Upload .csv file {i+1}", type="csv")
    uploaded_files.append(uploaded_file)

if all(uploaded_files):
    filenames = [file.name for file in uploaded_files]
    dataframes = [pd.read_csv(file) for file in uploaded_files]
    base_df = dataframes[0].copy()  # keep full original structure of file 1

    required_cols = ["spec_number", "cm_summary", "limits"]
    for df in dataframes:
        for col in required_cols:
            if col not in df.columns:
                st.error(f'Missing column "{col}" in one of the files.')
                st.stop()

    combined_columns = []
    for i, df in enumerate(dataframes):
        idx_cm_summary = df.columns.tolist().index("cm_summary")
        idx_limits = df.columns.tolist().index("limits")

        cl_columns = df.columns[idx_cm_summary: idx_cm_summary + 3].tolist()

        if i == 0:
            limit_columns = df.columns[idx_limits: idx_limits + 3].tolist()
            limit_data = df[limit_columns]
            limit_data.columns = ["Minimum_Limits1", "Typical_Limits1", "Maximum_Limits1"]
            columns = pd.concat([df[["spec_number", "spec_id_expansion"]], limit_data, df[cl_columns]], axis=1)
            columns.columns = ["spec_number", "spec_id_expansion",
                               "Minimum_Limits1", "Typical_Limits1", "Maximum_Limits1",
                               f"Minimum_CL{i+1}", f"Typical_CL{i+1}", f"Maximum_CL{i+1}"]
        else:
            columns = df[["spec_number", "spec_id_expansion"] + cl_columns]
            columns.columns = ["spec_number", "spec_id_expansion",
                               f"Minimum_CL{i+1}", f"Typical_CL{i+1}", f"Maximum_CL{i+1}"]

        combined_columns.append(columns)

    df_combined = combined_columns[0]
    for df in combined_columns[1:]:
        df_combined = pd.merge(df_combined, df, on=["spec_number", "spec_id_expansion"], how="outer")

    for col in df_combined.columns:
        if col.startswith("Minimum_") or col.startswith("Typical_") or col.startswith("Maximum_"):
            df_combined[col] = pd.to_numeric(df_combined[col], errors='coerce')

    limits_columns = ["Minimum_Limits1", "Typical_Limits1", "Maximum_Limits1"]
    cl_columns = [col for col in df_combined.columns if "CL" in col]
    df_combined = df_combined[["spec_number", "spec_id_expansion"] + limits_columns + cl_columns]

    file_keys_list = [set(zip(df["spec_number"], df["spec_id_expansion"])) for df in combined_columns]

    def check_file_presence(row):
        key = (row["spec_number"], row["spec_id_expansion"])
        found_in_files = [str(i + 1) for i, file_keys in enumerate(file_keys_list) if key in file_keys]
        if len(found_in_files) == num_files:
            return "Found in all files"
        elif len(found_in_files) == 1:
            return f"Only found in uploaded file {found_in_files[0]}"
        else:
            return f"Found in files: {', '.join(found_in_files)}"

    df_combined["File Presence"] = df_combined.apply(check_file_presence, axis=1)

    df_combined["spec_id_expansion_sort"] = df_combined["spec_id_expansion"].fillna("").apply(lambda x: (0, "") if x == "" else (1, x))
    df_combined = df_combined.sort_values(by=["spec_number", "spec_id_expansion_sort"]).drop(columns=["spec_id_expansion_sort"])

    # === NEW MERGE BACK TO ORIGINAL STRUCTURE ===
    result_columns = ["spec_number", "spec_id_expansion", "File Presence"]
    result_columns += [col for col in df_combined.columns if "CL" in col or "Limit" in col]

    df_combined["Pass or Fail"] = None
    df_combined["Why Failed"] = None
    result_columns += ["Pass or Fail", "Why Failed"]

    results_only_df = df_combined[result_columns]
    merged_output = pd.merge(base_df, results_only_df, on=["spec_number", "spec_id_expansion"], how="left")

    st.write("You can review your data below. Pass/Fail will be calculated in Excel.")
    st.dataframe(merged_output)

    # Visualization (line chart)
    unique_spec_numbers = df_combined["spec_number"].drop_duplicates().tolist()
    selected_spec_number = st.selectbox("Select Spec Number", unique_spec_numbers)

    filtered_data = df_combined[df_combined["spec_number"] == selected_spec_number]
    if not filtered_data.empty:
        cl_columns = [col for col in filtered_data.columns if "CL" in col]
        cl_data = filtered_data[["spec_number", "spec_id_expansion"] + cl_columns]
        cl_data_melted = cl_data.melt(id_vars=["spec_number", "spec_id_expansion"], var_name="CL Type", value_name="Value")
        cl_data_melted["File"] = cl_data_melted["CL Type"].str.extract(r'(\d)').astype(int)

        plt.figure(figsize=(10, 6))
        sns.lineplot(data=cl_data_melted, x="CL Type", y="Value", hue="File", style="File", markers=True)
        plt.axhline(y=filtered_data["Minimum_Limits1"].values[0], color='r', linestyle='--', label='Minimum Limit')
        plt.axhline(y=filtered_data["Typical_Limits1"].values[0], color='g', linestyle='--', label='Typical Limit')
        plt.axhline(y=filtered_data["Maximum_Limits1"].values[0], color='b', linestyle='--', label='Maximum Limit')
        plt.title(f"CL Comparison for Spec Number {selected_spec_number}")
        plt.xlabel("CL Type")
        plt.ylabel("Value")
        plt.legend(title="File")
        plt.xticks(rotation=45)
        st.pyplot(plt)

    # === WRITE TO EXCEL ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    ws.freeze_panes = "A2"
    ws.append(merged_output.columns.tolist())

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_font = Font(color="006100")
    red_font = Font(color="9C0006")

    for row_idx, row in merged_output.iterrows():
        ws.append(row.tolist())
        current_row = ws.max_row

        failures = []
        for i in range(1, num_files + 1):
            min_limit_val = row.get("Minimum_Limits1")
            max_limit_val = row.get("Maximum_Limits1")
            min_cl_val = row.get(f"Minimum_CL{i}")
            typ_cl_val = row.get(f"Typical_CL{i}")
            max_cl_val = row.get(f"Maximum_CL{i}")

            failed_min = False
            failed_max = False

            if min_limit_val is not None:
                if min_cl_val is not None and min_cl_val < min_limit_val:
                    failures.append(f"Minimum_CL{i} < Minimum_Limits1")
                if typ_cl_val is not None and typ_cl_val < min_limit_val:
                    failed_min = True

            if max_limit_val is not None:
                if max_cl_val is not None and max_cl_val > max_limit_val:
                    failures.append(f"Maximum_CL{i} > Maximum_Limits1")
                if typ_cl_val is not None and typ_cl_val > max_limit_val:
                    failed_max = True

            if typ_cl_val is not None and failed_min and failed_max:
                failures.append(f"Typical_CL{i} outside both limit bounds")

        pass_col = merged_output.columns.get_loc("Pass or Fail") + 1
        why_col = merged_output.columns.get_loc("Why Failed") + 1

        pass_val = "Pass" if not failures else "Fail"
        ws.cell(row=current_row, column=pass_col).value = pass_val
        ws.cell(row=current_row, column=why_col).value = ", ".join(failures)

        # Color formatting for that row
        for i in range(1, num_files + 1):
            min_col = merged_output.columns.get_loc(f"Minimum_CL{i}") + 1
            typ_col = merged_output.columns.get_loc(f"Typical_CL{i}") + 1
            max_col = merged_output.columns.get_loc(f"Maximum_CL{i}") + 1

            for col_idx, cl_val, limit_val, condition in [
                (min_col, row.get(f"Minimum_CL{i}"), row.get("Minimum_Limits1"), lambda cl, lim: cl >= lim),
                (max_col, row.get(f"Maximum_CL{i}"), row.get("Maximum_Limits1"), lambda cl, lim: cl <= lim)
            ]:
                cell = ws.cell(row=current_row, column=col_idx)
                if pd.isna(limit_val) or pd.isna(cl_val):
                    continue
                if condition(cl_val, limit_val):
                    cell.fill = green_fill
                    cell.font = green_font
                else:
                    cell.fill = red_fill
                    cell.font = red_font

            # Typical CL (check both bounds)
            typ_val = row.get(f"Typical_CL{i}")
            typ_cell = ws.cell(row=current_row, column=typ_col)
            fail_min = row.get("Minimum_Limits1") is not None and typ_val is not None and typ_val < row.get("Minimum_Limits1")
            fail_max = row.get("Maximum_Limits1") is not None and typ_val is not None and typ_val > row.get("Maximum_Limits1")
            if typ_val is not None and fail_min and fail_max:
                typ_cell.fill = red_fill
                typ_cell.font = red_font
            elif typ_val is not None:
                typ_cell.fill = green_fill
                typ_cell.font = green_font

    # Auto-width
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if isinstance(cell, Cell) and cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    final_output.name = "comparison_grouped.xlsx"

    st.download_button(
        label="Download Excel (Grouped with Original Columns)",
        data=final_output,
        file_name=final_output.name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )