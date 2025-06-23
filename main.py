import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
import matplotlib.pyplot as plt
import seaborn as sns

st.set_page_config(page_title="CL Comparison Tool", layout="centered")
st.title("📊 CL Comparison Tool")

num_files = st.selectbox("Select number of files to compare", [2, 3, 4], index=0)

uploaded_files = []
custom_names = []

for i in range(num_files):
    uploaded_file = st.file_uploader(f"Upload .csv file {i+1}", type="csv")
    uploaded_files.append(uploaded_file)
    
    if uploaded_file:
        custom_name = st.text_input(f"Enter a name for File {i+1}", value=f"File {i+1}")
    else:
        custom_name = f"File {i+1}"
    custom_names.append(custom_name)

if all(uploaded_files):
    dataframes = [pd.read_csv(file) for file in uploaded_files]
    base_df = dataframes[0].copy()

    required_cols = ["spec_number", "cm_summary", "limits", "spec_item_category", "spec_item_old_name"]
    for df in dataframes:
        for col in required_cols:
            if col not in df.columns:
                st.error(f'Missing column "{col}" in one of the files.')
                st.stop()

    combined_columns = []
    for i, df in enumerate(dataframes):
        idx_cm_summary = df.columns.tolist().index("cm_summary")
        idx_limits = df.columns.tolist().index("limits")

        cl_columns = df.columns[idx_cm_summary: idx_cm_summary + 3].tolist()

        if i == 0:
            limit_columns = df.columns[idx_limits: idx_limits + 3].tolist()
            limit_data = df[limit_columns]
            limit_data.columns = ["Minimum_Limits1", "Typical_Limits1", "Maximum_Limits1"]
            columns = pd.concat([df[["spec_number", "spec_id_expansion", "spec_item_category", "spec_item_old_name"]], limit_data, df[cl_columns]], axis=1)
            columns.columns = [
                "spec_number", "spec_id_expansion", "spec_item_category", "spec_item_old_name",
                "Minimum_Limits1", "Typical_Limits1", "Maximum_Limits1",
                f"Minimum_{custom_names[i]}", f"Typical_{custom_names[i]}", f"Maximum_{custom_names[i]}"
            ]
        else:
            columns = df[["spec_number", "spec_id_expansion", "spec_item_category", "spec_item_old_name"] + cl_columns]
            columns.columns = ["spec_number", "spec_id_expansion", "spec_item_category", "spec_item_old_name",
                               f"Minimum_{custom_names[i]}", f"Typical_{custom_names[i]}", f"Maximum_{custom_names[i]}"]

        combined_columns.append(columns)

    df_combined = combined_columns[0]
    for df in combined_columns[1:]:
        df_combined = pd.merge(df_combined, df, on=["spec_number", "spec_id_expansion", "spec_item_category", "spec_item_old_name"], how="outer")

    for col in df_combined.columns:
        if any(prefix in col for prefix in ["Minimum_", "Typical_", "Maximum_"]):
            df_combined[col] = pd.to_numeric(df_combined[col], errors='coerce')

    limits_columns = ["Minimum_Limits1", "Typical_Limits1", "Maximum_Limits1"]
    cl_columns = [col for name in custom_names for col in df_combined.columns if any(prefix in col for prefix in [f"Minimum_{name}", f"Typical_{name}", f"Maximum_{name}"])]
    df_combined = df_combined[["spec_number", "spec_id_expansion", "spec_item_category", "spec_item_old_name"] + limits_columns + cl_columns]

    file_keys_list = [set(zip(df["spec_number"], df["spec_id_expansion"], df["spec_item_category"], df["spec_item_old_name"])) for df in combined_columns]

    def check_file_presence(row):
        key = (row["spec_number"], row["spec_id_expansion"], row["spec_item_category"], row["spec_item_old_name"])
        found_in_files = [str(i + 1) for i, file_keys in enumerate(file_keys_list) if key in file_keys]
        if len(found_in_files) == num_files:
            return "Found in all files"
        elif len(found_in_files) == 1:
            return f"Only found in uploaded file {found_in_files[0]}"
        else:
            return f"Found in files: {', '.join(found_in_files)}"

    df_combined["File Presence"] = df_combined.apply(check_file_presence, axis=1)
    df_combined["spec_id_expansion_sort"] = df_combined["spec_id_expansion"].fillna("").apply(lambda x: (0, "") if x == "" else (1, x))
    df_combined = df_combined.sort_values(by=["spec_number", "spec_id_expansion_sort"]).drop(columns=["spec_id_expansion_sort"])

    result_columns = ["spec_number", "spec_id_expansion", "File Presence", "spec_item_category", "spec_item_old_name"] + limits_columns + cl_columns
    df_combined["Pass or Fail"] = None
    df_combined["Why Failed"] = None
    result_columns += ["Pass or Fail", "Why Failed"]

    results_only_df = df_combined[result_columns]
    merged_output = pd.merge(base_df, results_only_df, on=["spec_number", "spec_id_expansion", "spec_item_category", "spec_item_old_name"], how="left")

    columns_to_move = ["File Presence", "Minimum_Limits1", "Typical_Limits1", "Maximum_Limits1"]
    for name in custom_names:
        columns_to_move.extend([f"Minimum_{name}", f"Typical_{name}", f"Maximum_{name}"])
    columns_to_move.extend(["Pass or Fail", "Why Failed"])

    second_row = merged_output.iloc[0] if len(merged_output) > 0 else None
    vswr_column_name = None
    if second_row is not None:
        for col in merged_output.columns:
            if str(second_row[col]).strip().lower() == "vswr":
                vswr_column_name = col
                break

    if vswr_column_name:
        col_list = merged_output.columns.tolist()
        idx = col_list.index(vswr_column_name) + 1

        for col in columns_to_move:
            if col in col_list:
                col_list.remove(col)

        reordered_cols = col_list[:idx] + columns_to_move + col_list[idx:]
        merged_output = merged_output[reordered_cols]

    st.write("You can review your data below. Pass/Fail will be calculated in Excel.")
    st.dataframe(merged_output)
    st.header("Choose way of grouping to graph (Default is Spec Item Category)", divider=True)
    group_by_old_name = st.checkbox("Group by Spec Item Old Name", value=False)

    if group_by_old_name:
        unique_spec_item_names = df_combined["spec_item_old_name"].drop_duplicates().tolist()
        selected_spec_item_name = st.selectbox("Select Spec Item Old Name", unique_spec_item_names)
        filtered_data = df_combined[df_combined["spec_item_old_name"] == selected_spec_item_name]
    else:
        unique_spec_item_categorys = df_combined["spec_item_category"].drop_duplicates().tolist()
        selected_spec_item_category = st.selectbox("Select Spec Item Category", unique_spec_item_categorys)
        filtered_data = df_combined[df_combined["spec_item_category"] == selected_spec_item_category]

    unique_spec_numbers = filtered_data["spec_number"].dropna().unique()
    unique_spec_numbers.sort()
    selected_spec_numbers = st.multiselect("Filter by Spec Number(s)", unique_spec_numbers, default=unique_spec_numbers)
    filtered_data = filtered_data[filtered_data["spec_number"].isin(selected_spec_numbers)]

    st.header("Check boxes to Show/Hide Limits", divider=True)
    show_min_limit = st.checkbox("Show Minimum Limit", value=True)
    show_typ_limit = st.checkbox("Show Typical Limit", value=True)
    show_max_limit = st.checkbox("Show Maximum Limit", value=True)

    if not filtered_data.empty:
        cl_columns = [col for name in custom_names for col in filtered_data.columns if any(prefix in col for prefix in [f"Minimum_{name}", f"Typical_{name}", f"Maximum_{name}"])]
        cl_data = filtered_data[["spec_number", "spec_item_category", "spec_item_old_name"] + cl_columns]
        cl_data_melted = cl_data.melt(id_vars=["spec_number", "spec_item_category", "spec_item_old_name"], var_name="CL Type", value_name="Value")
        cl_data_melted["File"] = cl_data_melted["CL Type"].apply(
            lambda x: next((name for name in custom_names if name in x), "Unknown")
        )
        plt.figure(figsize=(20, 10))
        sns.lineplot(
            data=cl_data_melted,
            x="spec_number",
            y="Value",
            hue="File",
            style="File",
            markers=True,
            errorbar=None
        )
        if show_min_limit:
            plt.plot(filtered_data["spec_number"], filtered_data["Minimum_Limits1"], linestyle='--', color='purple', label='Minimum Limit')
        if show_typ_limit:
            plt.plot(filtered_data["spec_number"], filtered_data["Typical_Limits1"], linestyle='--', color='orange', label='Typical Limit')
        if show_max_limit:
            plt.plot(filtered_data["spec_number"], filtered_data["Maximum_Limits1"], linestyle='--', color='brown', label='Maximum Limit')
        plt.title(f"CL Comparison for {'Spec Item Old Name ' + selected_spec_item_name if group_by_old_name else 'Spec Item Category ' + selected_spec_item_category}")
        plt.xlabel("Spec Number")
        plt.ylabel("Value")
        plt.legend(title="File")
        plt.xticks(rotation=45)
        st.pyplot(plt)

        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', bbox_inches='tight')
        img_buffer.seek(0)
        st.download_button(
            label="Download This Graph as PNG",
            data=img_buffer,
            file_name="cl_comparison_graph.png",
            mime="image/png"
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    ws.freeze_panes = "A2"
    ws.append(merged_output.columns.tolist())
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_font = Font(color="006100")
    red_font = Font(color="9C0006")

    for row_idx, row in merged_output.iterrows():
        ws.append(row.tolist())
        current_row = ws.max_row
        failures = []
        for i in range(1, num_files + 1):
            name = custom_names[i - 1]
            min_limit_val = row.get("Minimum_Limits1")
            max_limit_val = row.get("Maximum_Limits1")
            min_cl_val = row.get(f"Minimum_{name}")
            typ_cl_val = row.get(f"Typical_{name}")
            max_cl_val = row.get(f"Maximum_{name}")
            failed_min = False
            failed_max = False
            if min_limit_val is not None:
                if min_cl_val is not None and min_cl_val < min_limit_val:
                    failures.append(f"Minimum_{name} < Minimum_Limits1")
                if typ_cl_val is not None and typ_cl_val < min_limit_val:
                    failed_min = True
            if max_limit_val is not None:
                if max_cl_val is not None and max_cl_val > max_limit_val:
                    failures.append(f"Maximum_{name} > Maximum_Limits1")
                if typ_cl_val is not None and typ_cl_val > max_limit_val:
                    failed_max = True
            if typ_cl_val is not None and failed_min and failed_max:
                failures.append(f"Typical_{name} outside both limit bounds")
        pass_col = merged_output.columns.get_loc("Pass or Fail") + 1
        why_col = merged_output.columns.get_loc("Why Failed") + 1

        pass_val = "Pass" if not failures else "Fail"
        ws.cell(row=current_row, column=pass_col).value = pass_val
        ws.cell(row=current_row, column=why_col).value = ", ".join(failures)
        pass_cell = ws.cell(row=current_row, column=pass_col)
        if pass_val == "Pass":
            pass_cell.fill = green_fill
            pass_cell.font = green_font
        else:
            pass_cell.fill = red_fill
            pass_cell.font = red_font
        for i in range(1, num_files + 1):
            name = custom_names[i - 1]
            min_col = merged_output.columns.get_loc(f"Minimum_{name}") + 1
            typ_col = merged_output.columns.get_loc(f"Typical_{name}") + 1
            max_col = merged_output.columns.get_loc(f"Maximum_{name}") + 1
            for col_idx, cl_val, limit_val, condition in [
                (min_col, row.get(f"Minimum_{name}"), row.get("Minimum_Limits1"), lambda cl, lim: cl >= lim),
                (max_col, row.get(f"Maximum_{name}"), row.get("Maximum_Limits1"), lambda cl, lim: cl <= lim)
            ]:
                cell = ws.cell(row=current_row, column=col_idx)
                if pd.isna(cl_val):
                    continue
                if pd.isna(limit_val):
                    cell.fill = green_fill
                    cell.font = green_font
                elif condition(cl_val, limit_val):
                    cell.fill = green_fill
                    cell.font = green_font
                else:
                    cell.fill = red_fill
                    cell.font = red_font
            typ_val = row.get(f"Typical_{name}")
            typ_cell = ws.cell(row=current_row, column=typ_col)
            fail_min = row.get("Minimum_Limits1") is not None and typ_val is not None and typ_val < row.get("Minimum_Limits1")
            fail_max = row.get("Maximum_Limits1") is not None and typ_val is not None and typ_val > row.get("Maximum_Limits1")
            if typ_val is not None:
                if fail_min and fail_max:
                    typ_cell.fill = red_fill
                    typ_cell.font = red_font
                else:
                    typ_cell.fill = green_fill
                    typ_cell.font = green_font

    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if isinstance(cell, Cell) and cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    final_output.name = "comparison_grouped.xlsx"
    st.download_button(
        label="Download Excel (Grouped with Original Columns)",
        data=final_output,
        file_name=final_output.name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
