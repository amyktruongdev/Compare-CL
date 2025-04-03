import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import os

st.set_page_config(page_title="CL Comparison Tool", layout="centered")

st.title("📊 CL1 vs CL2 Comparison Tool")

uploaded_file1 = st.file_uploader("Upload first .xlsx file (CL1)", type="xlsx")
uploaded_file2 = st.file_uploader("Upload second .xlsx file (CL2)", type="xlsx")

if uploaded_file1 and uploaded_file2:
    df_cl1 = pd.read_excel(uploaded_file1)
    df_cl2 = pd.read_excel(uploaded_file2)

    required_cols = ["spec_number", "cm_summary"]
    for col in required_cols:
        if col not in df_cl1.columns or col not in df_cl2.columns:
            st.error(f'Missing column "{col}" in one of the files.')
            st.stop()

    #Get next 2 columns after cm_summary
    idx_cl1 = df_cl1.columns.tolist().index("cm_summary")
    idx_cl2 = df_cl2.columns.tolist().index("cm_summary")

    cl1_columns = df_cl1[["spec_number", "spec_id_expansion"] + df_cl1.columns[idx_cl1: idx_cl1 + 3].tolist()]
    cl2_columns = df_cl2[["spec_number", "spec_id_expansion"] + df_cl2.columns[idx_cl2: idx_cl2 + 3].tolist()]

    cl1_columns.columns = ["spec_number", "spec_id_expansion", "Minimum_CL1", "Typical_CL1", "Maximum_CL1"]
    cl2_columns.columns = ["spec_number", "spec_id_expansion", "Minimum_CL2", "Typical_CL2", "Maximum_CL2"]

    df_combined = pd.merge(cl1_columns, cl2_columns, on=["spec_number", "spec_id_expansion"], how="outer")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        output_file = tmp.name

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_combined.to_excel(writer, sheet_name='Comparison', index=False)

    #Load and modify workbook
    wb = load_workbook(output_file)
    ws = wb["Comparison"]

    for spec_number in df_combined['spec_number'].dropna().unique():
        spec_group = df_combined[df_combined['spec_number'] == spec_number]
        if spec_group.empty:
            continue

        ws_chart = wb.create_sheet(title=f"{spec_number}_Comparison")

        header = []
        values = []

        for _, row in spec_group.iterrows():
            sid = row["spec_id_expansion"]
            header.extend([
                f"{sid}_CL1_Min", f"{sid}_CL1_Typ", f"{sid}_CL1_Max",
                f"{sid}_CL2_Min", f"{sid}_CL2_Typ", f"{sid}_CL2_Max"
            ])
            values.extend([
                float(row.get("Minimum_CL1", 0)),
                float(row.get("Typical_CL1", 0)),
                float(row.get("Maximum_CL1", 0)),
                float(row.get("Minimum_CL2", 0)),
                float(row.get("Typical_CL2", 0)),
                float(row.get("Maximum_CL2", 0))
            ])

        ws_chart.append(["Metric"] + header)
        ws_chart.append(["Values"] + values)

        chart = BarChart()
        chart.type = "col"
        chart.title = f"{spec_number} - CL1 & CL2 Comparison"
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Spec Grouped by Source"

        data = Reference(ws_chart, min_col=2, min_row=2, max_col=1 + len(values), max_row=2)
        categories = Reference(ws_chart, min_col=2, min_row=1, max_col=1 + len(values))
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)
        chart.series[0].graphicalProperties.solidFill = "808080"
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True

        ws_chart.add_chart(chart, "B5")

    wb.save(output_file)

    with open(output_file, "rb") as f:
        st.success("✅ File successfully generated!")
        st.download_button("📥 Download Result Excel File", f, file_name="side_by_side_comparison.xlsx")

